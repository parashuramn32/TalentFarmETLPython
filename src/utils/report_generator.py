"""Validation execution report (Assignment 3, Section 5).

  Execution Summary  : date/time, environment, code version, totals and counts of
                       passed / failed / skipped / BLOCKED
  Layer-wise Results : grouped by validation layer
  Detailed Results   : validation ID, expected, actual, VARIANCE, status, remarks
"""
from pathlib import Path
from datetime import datetime
import platform
import subprocess
import pandas as pd

from src.utils.logger import get_logger, run_log_path
from src.utils.config_loader import environment_name
from src.utils.result import PASS, FAIL, SKIPPED, BLOCKED

log = get_logger(__name__)
REPORT_DIR = Path(__file__).resolve().parents[2] / "reports"
REPORT_DIR.mkdir(exist_ok=True)

CODE_VERSION = "2.2.0"

STATUS_COLOR = {PASS: "C6EFCE", FAIL: "FFC7CE", BLOCKED: "FFD9B3", SKIPPED: "D9D9D9"}
STATUS_FONT = {PASS: "006100", FAIL: "9C0006", BLOCKED: "9C4500", SKIPPED: "595959"}

DETAIL_COLS = ["test_case_id", "layer", "description", "risk_ref", "source_object",
               "target_object", "expected", "actual", "variance", "variance_pct",
               "status", "severity", "remarks", "duration_sec", "executed_at"]


def _git_commit():
    try:
        out = subprocess.run(["git", "rev-parse", "--short", "HEAD"],
                             capture_output=True, text=True, timeout=5,
                             cwd=str(Path(__file__).resolve().parents[2]))
        return out.stdout.strip() or "n/a"
    except Exception:
        return "n/a"


def _frame(results):
    if not results:
        return pd.DataFrame(columns=DETAIL_COLS)
    return pd.DataFrame([r.to_dict() for r in results])


def summarise(results, business_date=None):
    df = _frame(results)
    total = len(df)
    counts = {s: int((df["status"] == s).sum()) if total else 0
              for s in (PASS, FAIL, SKIPPED, BLOCKED)}
    executed = total - counts[SKIPPED] - counts[BLOCKED]
    return {
        "execution_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "business_date_validated": business_date or "n/a",
        "environment": environment_name(),
        "code_version": f"{CODE_VERSION} (commit {_git_commit()})",
        "python_version": platform.python_version(),
        "executed_by": platform.node(),
        "total_validations": total,
        "passed": counts[PASS],
        "failed": counts[FAIL],
        "skipped": counts[SKIPPED],
        "blocked": counts[BLOCKED],
        "pass_rate_pct": round(100 * counts[PASS] / executed, 1) if executed else 0.0,
        "duration_sec": round(sum(r.duration_sec for r in results), 2),
        "log_file": run_log_path().name,
    }


def layer_summary(results):
    df = _frame(results)
    if df.empty:
        return pd.DataFrame()
    piv = (df.pivot_table(index="layer", columns="status", values="test_case_id",
                          aggfunc="count", fill_value=0).reset_index())
    for s in (PASS, FAIL, SKIPPED, BLOCKED):
        if s not in piv.columns:
            piv[s] = 0
    piv["TOTAL"] = piv[[PASS, FAIL, SKIPPED, BLOCKED]].sum(axis=1)
    executed = piv[PASS] + piv[FAIL]
    piv["PASS_RATE_PCT"] = (100 * piv[PASS] / executed.replace(0, pd.NA)).round(1).fillna(0)
    return piv[["layer", "TOTAL", PASS, FAIL, SKIPPED, BLOCKED, "PASS_RATE_PCT"]]


def _style(xl):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in xl.book.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
            cell.fill = PatternFill("solid", start_color="1F3864")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        headers = [c.value for c in sheet[1]]
        if "status" in headers:
            idx = headers.index("status") + 1
            for row in range(2, sheet.max_row + 1):
                c = sheet.cell(row, idx)
                if str(c.value) in STATUS_COLOR:
                    c.fill = PatternFill("solid", start_color=STATUS_COLOR[str(c.value)])
                    c.font = Font(bold=True, color=STATUS_FONT[str(c.value)],
                                  size=9, name="Arial")
                    c.alignment = Alignment(horizontal="center")
        for i, col in enumerate(headers, 1):
            width = (46 if col in ("description", "remarks") else
                     26 if col in ("source_object", "target_object", "expected",
                                   "actual", "variance") else
                     30 if col == "Value" else 15)
            sheet.column_dimensions[get_column_letter(i)].width = width
        sheet.freeze_panes = "A2"


def to_excel(results, filename=None, business_date=None):
    df = _frame(results)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = REPORT_DIR / (filename or f"validation_report_{stamp}.xlsx")
    summary = summarise(results, business_date)
    detail = df[[c for c in DETAIL_COLS if c in df.columns]]

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        (pd.DataFrame(list(summary.items()), columns=["Metric", "Value"])
           .to_excel(xl, sheet_name="Execution Summary", index=False))
        ls = layer_summary(results)
        if not ls.empty:
            ls.to_excel(xl, sheet_name="Layer-wise Results", index=False)
        detail.to_excel(xl, sheet_name="Detailed Results", index=False)
        if not detail.empty:
            fails = detail[detail["status"] == FAIL]
            if not fails.empty:
                fails.to_excel(xl, sheet_name="Failures", index=False)
            blocked = detail[detail["status"] == BLOCKED]
            if not blocked.empty:
                blocked.to_excel(xl, sheet_name="Blocked", index=False)
        _style(xl)

    log.info("Excel validation report written: %s", path)
    return path


def to_html(results, filename=None, business_date=None):
    df = _frame(results)
    s = summarise(results, business_date)
    ls = layer_summary(results)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = REPORT_DIR / (filename or f"validation_report_{stamp}.html")

    meta_rows = "".join(
        f"<tr><th>{k.replace('_', ' ').title()}</th><td>{v}</td></tr>"
        for k, v in s.items()
        if k not in ("passed", "failed", "skipped", "blocked",
                     "total_validations", "pass_rate_pct"))
    layer_rows = "".join("<tr>" + "".join(f"<td>{v}</td>" for v in row) + "</tr>"
                         for row in ls.itertuples(index=False)) if not ls.empty else ""
    detail_rows = "".join(
        f"<tr class='{str(r.get('status','')).lower()}'><td>{r.get('test_case_id','')}</td>"
        f"<td>{r.get('layer','')}</td><td>{r.get('description','')}</td>"
        f"<td>{r.get('risk_ref','')}</td><td>{r.get('expected','')}</td>"
        f"<td>{r.get('actual','')}</td><td>{r.get('variance','')}</td>"
        f"<td class='st'>{r.get('status','')}</td>"
        f"<td>{r.get('severity','')}</td><td>{r.get('remarks','')}</td></tr>"
        for _, r in df.iterrows())

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Validation Execution Report</title><style>
body{{font-family:Arial,Helvetica,sans-serif;margin:24px;color:#222}}
h1{{color:#1F3864;margin-bottom:2px}} h2{{color:#1F3864;margin-top:28px}}
.sub{{color:#666;margin-bottom:18px}}
.cards{{display:flex;gap:12px;flex-wrap:wrap;margin:18px 0}}
.card{{border:1px solid #ddd;border-radius:8px;padding:12px 18px;min-width:110px}}
.card .n{{font-size:26px;font-weight:bold}} .card .l{{font-size:12px;color:#666}}
.pass .n{{color:#137333}} .fail .n{{color:#c00}}
.blocked .n{{color:#b35c00}} .skip .n{{color:#666}}
table{{border-collapse:collapse;width:100%;font-size:12px;margin-bottom:10px}}
th{{background:#1F3864;color:#fff;padding:8px;text-align:left}}
td{{border:1px solid #ddd;padding:6px;vertical-align:top}}
table.meta th{{width:220px;background:#EAF1FA;color:#1F3864}}
tr.pass td.st{{background:#C6EFCE;color:#006100;font-weight:bold;text-align:center}}
tr.fail td.st{{background:#FFC7CE;color:#9C0006;font-weight:bold;text-align:center}}
tr.blocked td.st{{background:#FFD9B3;color:#9C4500;font-weight:bold;text-align:center}}
tr.skipped td.st{{background:#D9D9D9;color:#595959;font-weight:bold;text-align:center}}
</style></head><body>
<h1>Validation Execution Report</h1>
<div class="sub">Python QA Automation &middot; Financial Services Sales Data Pipeline</div>
<h2>Execution Summary</h2>
<table class="meta">{meta_rows}</table>
<div class="cards">
<div class="card"><div class="n">{s['total_validations']}</div><div class="l">Total</div></div>
<div class="card pass"><div class="n">{s['passed']}</div><div class="l">Passed</div></div>
<div class="card fail"><div class="n">{s['failed']}</div><div class="l">Failed</div></div>
<div class="card skip"><div class="n">{s['skipped']}</div><div class="l">Skipped</div></div>
<div class="card blocked"><div class="n">{s['blocked']}</div><div class="l">Blocked</div></div>
<div class="card"><div class="n">{s['pass_rate_pct']}%</div><div class="l">Pass Rate</div></div>
</div>
<h2>Layer-wise Results</h2>
<table><thead><tr><th>Layer</th><th>Total</th><th>Passed</th><th>Failed</th>
<th>Skipped</th><th>Blocked</th><th>Pass Rate %</th></tr></thead>
<tbody>{layer_rows}</tbody></table>
<h2>Detailed Results</h2>
<table><thead><tr><th>Validation ID</th><th>Layer</th><th>Description</th><th>Risk Ref</th>
<th>Expected</th><th>Actual</th><th>Variance</th><th>Status</th><th>Severity</th>
<th>Remarks</th></tr></thead><tbody>{detail_rows}</tbody></table>
</body></html>"""
    path.write_text(html, encoding="utf-8")
    log.info("HTML validation report written: %s", path)
    return path


def print_console(results, business_date=None):
    s = summarise(results, business_date)
    print("\n" + "=" * 82)
    print("  VALIDATION EXECUTION SUMMARY")
    print("=" * 82)
    print(f"  Executed at   : {s['execution_datetime']}")
    print(f"  Business date : {s['business_date_validated']}")
    print(f"  Environment   : {s['environment']}")
    print(f"  Code version  : {s['code_version']}")
    print(f"  Log file      : {s['log_file']}")
    print("-" * 82)
    print(f"  Total: {s['total_validations']}   Passed: {s['passed']}   "
          f"Failed: {s['failed']}   Skipped: {s['skipped']}   Blocked: {s['blocked']}   "
          f"Pass rate: {s['pass_rate_pct']}%")
    print("-" * 82)
    ls = layer_summary(results)
    if not ls.empty:
        print(f"  {'LAYER':<26}{'TOT':>5}{'PASS':>6}{'FAIL':>6}{'SKIP':>6}{'BLOCK':>7}{'RATE':>8}")
        for row in ls.itertuples(index=False):
            print(f"  {row.layer:<26}{row.TOTAL:>5}{getattr(row, PASS):>6}"
                  f"{getattr(row, FAIL):>6}{getattr(row, SKIPPED):>6}"
                  f"{getattr(row, BLOCKED):>7}{row.PASS_RATE_PCT:>7}%")
        print("-" * 82)
    for r in results:
        icon = {PASS: "[PASS ]", FAIL: "[FAIL ]",
                BLOCKED: "[BLOCK]", SKIPPED: "[SKIP ]"}.get(r.status, "[     ]")
        print(f"  {icon} {r.test_case_id:<9} {r.layer:<24} {r.description[:40]}")
        if r.status in (FAIL, BLOCKED):
            print(f"           -> {r.message}")
    print("=" * 82 + "\n")
