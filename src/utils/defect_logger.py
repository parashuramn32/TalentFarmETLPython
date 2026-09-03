"""Defect log generation (Assignment 3, Section 6).

Converts FAILED validations into a defect log with the twelve required fields:
  Defect ID | Layer | Title | Description | Expected | Actual | Source Object |
  Target Object | Severity | Evidence | Likely Root Cause | Business Impact

Severity is normalised to the three-level scale in Section 8 via `severity_map`.
Only genuine FAIL results become defects - BLOCKED and SKIPPED do not.
"""
from pathlib import Path
from datetime import datetime
import pandas as pd

from src.utils.config_loader import load_rules, environment_name
from src.utils.logger import get_logger, run_log_path
from src.utils.result import FAIL

log = get_logger(__name__)
REPORT_DIR = Path(__file__).resolve().parents[2] / "reports"
REPORT_DIR.mkdir(exist_ok=True)

LAYER_CODE = {
    "Source-to-Staging": "STG",
    "Data Quality": "DQ",
    "Staging-to-Data Mart": "DM",
    "Aggregate Model": "AGG",
    "Data Mart-to-Reporting": "RPT",
    "Regression": "REG",
}

ROOT_CAUSE = {
    "R-SS-01": "The load filter or incremental window in the source-to-staging job is excluding "
               "valid records, or the job did not process the full source set.",
    "R-SS-02": "The staging load is not idempotent - a re-run inserted the same natural key "
               "twice because the staging table has no primary key constraint.",
    "R-SS-03": "The status filter in the staging load is missing or uses the wrong column or "
               "value, so excluded transaction states are being loaded.",
    "R-SS-04": "The distributor_txn_id to transaction_id rename is dropping or truncating "
               "values during the load mapping.",
    "R-SS-05": "Mandatory-field enforcement is absent in the staging load; nulls from source "
               "pass through unchecked.",
    "R-SS-06": "The API extraction stops early - the pagination loop does not iterate until "
               "total_records is reached.",
    "R-SS-07": "The source structure changed without a corresponding change to the load mapping.",
    "R-SS-08": "Source data types are not validated or cast during extraction.",
    "R-SS-09": "Column mapping in the staging load is transposed or applies an unintended "
               "transformation.",
    "R-SS-10": "The load is picking up a file or partition for the wrong business date.",
    "R-SS-11": "Audit columns are not populated by the load job.",
    "R-SS-13": "Commission calculation in the source application allows a value greater than "
               "the sale amount.",
    "R-SS-14": "The API service is down, misconfigured, or the endpoint has moved.",
    "R-SS-15": "API authentication is not enforced on the endpoint.",
    "R-SS-16": "The API contract changed and fields were renamed or removed.",
    "R-SS-17": "Date filtering is applied after pagination, or the API ignores the date "
               "parameters.",
    "R-DQ-01": "The master data load did not complete, or filtered rows that should be retained.",
    "R-DQ-02": "No format validation on mobile capture in the source application.",
    "R-DQ-03": "Free-text state entry with no controlled vocabulary; the normalisation "
               "dictionary does not cover all observed variants.",
    "R-DQ-04": "Reference data is stale relative to transactional data - new codes are in use "
               "before the master is refreshed.",
    "R-DQ-05": "Customer records are created after the transaction, or the customer master load "
               "runs before the transactional load.",
    "R-DQ-06": "No KYC gate in the transformation - rejected customers are not excluded from "
               "revenue recognition.",
    "R-DQ-07": "No validation on date of birth at capture.",
    "R-DQ-08": "Product master maintained manually without a domain constraint.",
    "R-DQ-09": "Region reference maintained manually without a domain constraint.",
    "R-DQ-10": "Conditional mandatory rules are not enforced by product type at capture.",
    "R-DM-01": "The product harmonisation join is failing or the mapping table is out of date.",
    "R-DM-02": "ULIP is classified by name pattern rather than by the product master, so it "
               "falls into the mutual fund branch of the transformation.",
    "R-DM-03": "The transformation copies the raw product label instead of joining to the "
               "master for the standardised name and category.",
    "R-DM-04": "Channel derivation uses an unmapped or free-text source value.",
    "R-DM-05": "The branch/region join fails or falls through to a default before all mapping "
               "options are attempted.",
    "R-DM-06": "State normalisation runs before or without the full variant dictionary, so the "
               "region lookup misses.",
    "R-DM-08": "Name cleansing (trim, salutation strip, title case) is not applied, or is "
               "applied in the wrong order, in the transformation.",
    "R-DM-09": "The customer master join is not applied for this channel.",
    "R-DM-10": "The net amount formula in the transformation does not match the agreed "
               "definition of gross minus discount.",
    "R-DM-11": "No active_flag check in the transformation, or the source application allows "
               "sales against discontinued products.",
    "R-DM-12": "The transformation is filtering, duplicating or failing to load part of the "
               "valid staging set.",
    "R-DM-13": "Null-customer transactions are not excluded or flagged before revenue "
               "aggregation.",
    "R-DM-14": "Reference data gaps cause the region derivation to fall through to the "
               "UNKNOWN_REGION default.",
    "R-AG-01": "The aggregate build is running against a different filter set or a stale "
               "snapshot of the transaction table.",
    "R-AG-02": "The GROUP BY grain in the summary build does not match the intended grain.",
    "R-AG-03": "The average is computed on gross instead of net, or divides without guarding "
               "against a zero transaction count.",
    "R-AG-04": "The executive summary is built from a summary table rather than the transaction "
               "detail, so it inherits an upstream error.",
    "R-AG-05": "The ranking query orders on the wrong measure or does not break ties "
               "deterministically.",
    "R-AG-06": "Summary models are refreshed independently and are out of step with each other.",
    "R-RP-01": "The dashboard query applies a different filter than the reporting view, or the "
               "page is caching a previous result.",
    "R-RP-02": "The filter parameter is not passed through to the underlying query.",
    "R-RP-03": "The date range boundary in the report query is inclusive or exclusive in the "
               "wrong direction.",
    "R-RP-04": "The reporting refresh did not run after the latest data mart load.",
    "R-RP-05": "The view definition adds a filter or join that was not present in the data mart "
               "table.",
    "R-RP-06": "Reports are sourced from different tables refreshed at different times.",
    "R-RG-01": "The fix addressed the reported symptom but changed shared transformation logic.",
    "R-RG-02": "The new load introduced data that the existing logic does not handle.",
    "R-RG-03": "The validation suite has a coverage gap for this defect pattern.",
}

DEFECT_COLS = ["Defect ID", "Layer", "Title", "Description", "Expected Value",
               "Actual Value", "Source Object", "Target Object", "Severity",
               "Evidence", "Likely Root Cause", "Business Impact",
               "Validation ID", "Risk Reference", "Detected On", "Environment", "Status"]


def _evidence(result):
    parts = [f"Validation {result.test_case_id} failed during the run recorded in "
             f"{run_log_path().name}."]
    if result.message:
        parts.append(f"Result: {result.message}")
    if result.failed_sample:
        parts.append(f"Sample offending rows: {str(result.failed_sample)[:400]}")
    parts.append(f"Full evidence: reports/evidence/{result.test_case_id}_FAIL.json")
    return " | ".join(parts)


def build_defects(results, business_date=None):
    rules = load_rules()
    sev_map = rules.get("severity_map", {})
    impact_map = rules.get("business_impact", {})
    env = environment_name()
    detected = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    defects, counters = [], {}
    for r in results:
        if r.status != FAIL:
            continue
        code = LAYER_CODE.get(r.layer, "GEN")
        counters[code] = counters.get(code, 0) + 1
        risk = (r.risk_ref or "").split()[0] if r.risk_ref else ""
        defects.append({
            "Defect ID": f"DEF-{code}-{counters[code]:03d}",
            "Layer": r.layer,
            "Title": r.description,
            "Description": (f"Validation {r.test_case_id} compared "
                            f"{r.source_object or 'source'} against "
                            f"{r.target_object or 'target'} for business date "
                            f"{business_date or 'n/a'} and detected a mismatch. {r.message}"),
            "Expected Value": r.expected,
            "Actual Value": r.actual,
            "Source Object": r.source_object,
            "Target Object": r.target_object,
            "Severity": sev_map.get(r.severity, r.severity),
            "Evidence": _evidence(r),
            "Likely Root Cause": ROOT_CAUSE.get(
                risk, "Root cause to be confirmed with the development team during triage."),
            "Business Impact": impact_map.get(
                risk, "Reported figures may be inaccurate for this metric."),
            "Validation ID": r.test_case_id,
            "Risk Reference": r.risk_ref,
            "Detected On": detected,
            "Environment": env,
            "Status": "Open",
        })
    log.info("Defect log built: %s defect(s) from %s validation result(s)",
             len(defects), len(results))
    return defects


def severity_breakdown(defects):
    out = {"High": 0, "Medium": 0, "Low": 0}
    for d in defects:
        out[d["Severity"]] = out.get(d["Severity"], 0) + 1
    return out


def to_excel(results, filename=None, business_date=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    defects = build_defects(results, business_date)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = REPORT_DIR / (filename or f"defect_log_{stamp}.xlsx")
    df = pd.DataFrame(defects, columns=DEFECT_COLS) if defects else \
        pd.DataFrame(columns=DEFECT_COLS)

    breakdown = severity_breakdown(defects)
    summary = {
        "Business date validated": business_date or "n/a",
        "Environment": environment_name(),
        "Defect log generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Total defects raised": len(defects),
        "High severity": breakdown.get("High", 0),
        "Medium severity": breakdown.get("Medium", 0),
        "Low severity": breakdown.get("Low", 0),
        "Source log file": run_log_path().name,
    }

    sev_fill = {"High": "FFC7CE", "Medium": "FFEB9C", "Low": "C6EFCE"}
    sev_font = {"High": "9C0006", "Medium": "9C6500", "Low": "006100"}

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        (pd.DataFrame(list(summary.items()), columns=["Metric", "Value"])
           .to_excel(xl, sheet_name="Defect Summary", index=False))
        df.to_excel(xl, sheet_name="Defect Log", index=False)
        if defects:
            (df.groupby(["Layer", "Severity"]).size().unstack(fill_value=0).reset_index()
               .to_excel(xl, sheet_name="By Layer", index=False))

        thin = Side(style="thin", color="A6A6A6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for sheet in xl.book.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
                cell.fill = PatternFill("solid", start_color="1F3864")
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border = border
            headers = [c.value for c in sheet[1]]
            for row in range(2, sheet.max_row + 1):
                for c in sheet[row]:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    c.font = Font(size=9, name="Arial")
                    c.border = border
            if "Severity" in headers:
                idx = headers.index("Severity") + 1
                for row in range(2, sheet.max_row + 1):
                    c = sheet.cell(row, idx)
                    if str(c.value) in sev_fill:
                        c.fill = PatternFill("solid", start_color=sev_fill[str(c.value)])
                        c.font = Font(bold=True, color=sev_font[str(c.value)],
                                      size=9, name="Arial")
                        c.alignment = Alignment(horizontal="center")
            for i, col in enumerate(headers, 1):
                width = (58 if col in ("Description", "Evidence", "Likely Root Cause",
                                       "Business Impact") else
                         38 if col == "Title" else
                         26 if col in ("Source Object", "Target Object", "Expected Value",
                                       "Actual Value", "Value") else 16)
                sheet.column_dimensions[get_column_letter(i)].width = width
            sheet.freeze_panes = "A2"

    log.info("Defect log written: %s (%s defects)", path, len(defects))
    return path, defects
