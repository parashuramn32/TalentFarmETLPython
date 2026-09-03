"""
Final QA Summary generator (Assignment 3, Section 10).

Reads the validation report and defect log produced by a real execution run and
generates the Final QA Summary document. Sections that require QA judgement are
pre-filled with a data-driven draft that you review and adjust.

Place this file at:  src/utils/qa_summary.py

Usage
-----
    # after a real run
    python -m src.utils.qa_summary --auto

    # or point at specific files
    python -m src.utils.qa_summary \
        --report reports/validation_report_20260904_101500.xlsx \
        --defects reports/defect_log_20260904_101500.xlsx
"""
import argparse
import glob
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

REPORT_DIR = Path(__file__).resolve().parents[2] / "reports"

# Sign-off thresholds. Tune these with your mentor if the programme defines its own.
THRESHOLDS = {
    "pass_rate_go": 95.0,            # >= this and no high-severity defects -> GO
    "pass_rate_conditional": 85.0,   # >= this -> GO WITH CONDITIONS
    "max_high_for_go": 0,
    "max_high_for_conditional": 3,
}

LAYER_ORDER = ["Source-to-Staging", "Data Quality", "Staging-to-Data Mart",
               "Aggregate Model", "Data Mart-to-Reporting", "Regression"]


# --------------------------------------------------------------- loading
def _latest(pattern):
    files = sorted(glob.glob(str(REPORT_DIR / pattern)), key=os.path.getmtime)
    return files[-1] if files else None


def load_inputs(report_path=None, defect_path=None):
    report_path = report_path or _latest("validation_report_*.xlsx")
    defect_path = defect_path or _latest("defect_log_*.xlsx")
    if not report_path:
        sys.exit("No validation report found. Run: python -m src.main --date <d> --regression")

    summary = pd.read_excel(report_path, sheet_name="Execution Summary")
    meta = dict(zip(summary["Metric"], summary["Value"]))
    detail = pd.read_excel(report_path, sheet_name="Detailed Results")
    try:
        layers = pd.read_excel(report_path, sheet_name="Layer-wise Results")
    except ValueError:
        layers = pd.DataFrame()

    defects = pd.DataFrame()
    if defect_path and Path(defect_path).exists():
        try:
            defects = pd.read_excel(defect_path, sheet_name="Defect Log")
        except ValueError:
            pass

    return {"meta": meta, "detail": detail, "layers": layers, "defects": defects,
            "report_path": report_path, "defect_path": defect_path}


# --------------------------------------------------------------- analysis
def assess(meta, defects):
    """Derive the recommendation from the run. Review before submitting."""
    pass_rate = float(meta.get("pass_rate_pct", 0) or 0)
    blocked = int(meta.get("blocked", 0) or 0)
    high = 0
    if not defects.empty and "Severity" in defects.columns:
        high = int((defects["Severity"] == "High").sum())

    if blocked > 0:
        verdict = "NO-GO — INCOMPLETE"
        rationale = (f"{blocked} validation(s) could not be executed. The run does not "
                     f"provide full coverage, so no sign-off can be given until the "
                     f"environment issues are resolved and the suite is re-run.")
    elif pass_rate >= THRESHOLDS["pass_rate_go"] and high <= THRESHOLDS["max_high_for_go"]:
        verdict = "GO"
        rationale = (f"Pass rate of {pass_rate}% with no high-severity defects. The pipeline "
                     f"is behaving as specified for the validated business date.")
    elif (pass_rate >= THRESHOLDS["pass_rate_conditional"]
          and high <= THRESHOLDS["max_high_for_conditional"]):
        verdict = "GO WITH CONDITIONS"
        rationale = (f"Pass rate of {pass_rate}% with {high} high-severity defect(s). "
                     f"Reporting may be released once the high-severity defects are fixed "
                     f"and the affected validations are re-executed.")
    else:
        verdict = "NO-GO"
        rationale = (f"Pass rate of {pass_rate}% with {high} high-severity defect(s). "
                     f"Reported figures cannot be relied upon until these are resolved.")
    return verdict, rationale, high, blocked


def top_findings(defects, n=5):
    if defects.empty:
        return []
    d = defects.copy()
    order = {"High": 0, "Medium": 1, "Low": 2}
    d["_o"] = d["Severity"].map(order).fillna(3)
    d = d.sort_values("_o").head(n)
    return d.to_dict("records")


def themes(defects):
    """Group defects into recurring themes so the summary reads analytically."""
    if defects.empty:
        return []
    out = []
    for layer, grp in defects.groupby("Layer"):
        sev = grp["Severity"].value_counts().to_dict()
        out.append({
            "layer": layer,
            "count": len(grp),
            "severity": ", ".join(f"{k}: {v}" for k, v in sev.items()),
            "examples": ", ".join(grp["Defect ID"].head(4).tolist()),
        })
    return sorted(out, key=lambda x: -x["count"])


# --------------------------------------------------------------- output
def build(data, out_path=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    meta, detail, layers, defects = (data["meta"], data["detail"],
                                     data["layers"], data["defects"])
    verdict, rationale, high, blocked = assess(meta, defects)

    HDR, ALT, BAND = "1F3864", "EAF1FA", "DBE5F1"
    thin = Side(style="thin", color="A6A6A6")
    BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
    F = "Arial"

    wb = Workbook()

    def sheet(title):
        ws = wb.create_sheet(title)
        return ws

    def head(ws, row, cols, widths):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row, c, h)
            cell.font = Font(name=F, bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", start_color=HDR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORD
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    def body(ws, start, rows, bold_first=False):
        for r, row in enumerate(rows, start):
            for c, v in enumerate(row, 1):
                cell = ws.cell(r, c, v)
                cell.font = Font(name=F, size=9, bold=(bold_first and c == 1))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = BORD
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", start_color=ALT)

    def title(ws, text, span=4):
        ws["A1"] = text
        ws["A1"].font = Font(name=F, bold=True, size=14, color=HDR)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)

    # ---------------- 1. Executive Summary ----------------
    ws = wb.active
    ws.title = "1. Executive Summary"
    title(ws, "Final QA Summary — Financial Services Sales Data Pipeline")
    ws["A2"] = "Submission 3 · Test Execution and Defect Reporting"
    ws["A2"].font = Font(name=F, size=10, italic=True, color="595959")

    rows = [
        ["Business date validated", meta.get("business_date_validated", "")],
        ["Environment", meta.get("environment", "")],
        ["Execution date/time", meta.get("execution_datetime", "")],
        ["Code version", meta.get("code_version", "")],
        ["Total validations executed", meta.get("total_validations", "")],
        ["Passed", meta.get("passed", "")],
        ["Failed", meta.get("failed", "")],
        ["Skipped", meta.get("skipped", "")],
        ["Blocked", meta.get("blocked", "")],
        ["Pass rate", f"{meta.get('pass_rate_pct', '')}%"],
        ["Defects raised", 0 if defects.empty else len(defects)],
        ["High-severity defects", high],
        ["", ""],
        ["QA RECOMMENDATION", verdict],
        ["Rationale", rationale],
    ]
    head(ws, 4, ["Metric", "Value"], [34, 96])
    body(ws, 5, rows, bold_first=True)
    for r in range(5, 5 + len(rows)):
        if ws.cell(r, 1).value == "QA RECOMMENDATION":
            for c in (1, 2):
                ws.cell(r, c).fill = PatternFill("solid", start_color=BAND)
                ws.cell(r, c).font = Font(name=F, bold=True, size=11, color=HDR)

    # ---------------- 2. Layer-wise Results ----------------
    ws = sheet("2. Layer-wise Results")
    title(ws, "Validation results by pipeline layer", 7)
    if not layers.empty:
        head(ws, 3, list(layers.columns), [26, 10, 10, 10, 10, 10, 14])
        body(ws, 4, layers.values.tolist())
    else:
        ws["A3"] = "Layer-wise sheet not present in the validation report."

    # ---------------- 3. Key Findings ----------------
    ws = sheet("3. Key Findings")
    title(ws, "Key findings — highest severity first", 6)
    head(ws, 3, ["Defect ID", "Layer", "Title", "Severity",
                 "Business Impact", "Likely Root Cause"], [14, 22, 40, 11, 52, 52])
    rows = [[d.get("Defect ID"), d.get("Layer"), d.get("Title"), d.get("Severity"),
             d.get("Business Impact"), d.get("Likely Root Cause")]
            for d in top_findings(defects, 10)]
    body(ws, 4, rows or [["", "", "No defects raised in this run.", "", "", ""]])

    # ---------------- 4. Defect Themes ----------------
    ws = sheet("4. Defect Themes")
    title(ws, "Recurring themes by layer", 4)
    head(ws, 3, ["Layer", "Defect count", "Severity split", "Example defect IDs"],
         [28, 14, 26, 40])
    th = themes(defects)
    body(ws, 4, [[t["layer"], t["count"], t["severity"], t["examples"]] for t in th]
         or [["", "", "No defects raised.", ""]])

    note_r = 5 + max(len(th), 1)
    ws.cell(note_r, 1, "Interpretation (review and edit before submitting)").font = Font(
        name=F, bold=True, size=11, color=HDR)
    interp = [
        "Layers with the highest defect counts indicate where the pipeline logic is weakest.",
        "A failure at an upstream layer usually explains failures downstream - fix in execution order.",
        "Repeated root causes across defects point to a single underlying code change rather than many.",
    ]
    for i, t in enumerate(interp):
        ws.cell(note_r + 1 + i, 1, t).font = Font(name=F, size=9)
        ws.merge_cells(start_row=note_r + 1 + i, start_column=1,
                       end_row=note_r + 1 + i, end_column=4)

    # ---------------- 5. Risk Assessment ----------------
    ws = sheet("5. Risk Assessment")
    title(ws, "Residual risk and coverage", 3)
    head(ws, 3, ["Area", "Assessment", "Action required"], [30, 62, 46])
    risk_rows = [
        ["Execution coverage",
         f"{meta.get('total_validations','')} validations executed; "
         f"{meta.get('skipped',0)} skipped, {meta.get('blocked',0)} blocked.",
         "Re-run any blocked checks before sign-off." if blocked else "None."],
        ["Data accuracy",
         f"Pass rate {meta.get('pass_rate_pct','')}% with {high} high-severity defect(s).",
         "Fix high-severity defects and re-validate." if high else "Monitor on next load."],
        ["Reporting reliability",
         "Derived from the reporting-layer results in sheet 2.",
         "Confirm dashboards reconcile after fixes."],
        ["Reference data freshness",
         "Orphan reference checks (MST-V10..V13) indicate whether master data lags transactions.",
         "Align master refresh ahead of the transactional load if orphans persist."],
        ["Regression confidence",
         "Suite verified executable end-to-end; seeded-defect detection demonstrated.",
         "Re-run the full suite after each fix."],
    ]
    body(ws, 4, risk_rows, bold_first=True)

    # ---------------- 6. Recommendations ----------------
    ws = sheet("6. Recommendations")
    title(ws, "Recommendations and next steps", 3)
    head(ws, 3, ["#", "Recommendation", "Owner / Priority"], [5, 92, 26])
    recs = []
    if blocked:
        recs.append("Resolve the environment/access issues causing blocked validations and "
                    "re-execute the suite to obtain full coverage.")
    if high:
        recs.append(f"Prioritise the {high} high-severity defect(s) - these distort executive "
                    f"metrics or represent a compliance breach.")
    recs += [
        "Fix defects in execution order (source-to-staging first); downstream failures are "
        "often symptoms of an upstream cause.",
        "Re-run the full suite after each fix to confirm no regression is introduced.",
        "Schedule the suite against each daily load so defects are caught at source rather "
        "than discovered in executive reporting.",
        "Extend automated coverage to any layer where defects were found manually.",
    ]
    body(ws, 4, [[i, r, ""] for i, r in enumerate(recs, 1)])

    # ---------------- 7. Sign-off ----------------
    ws = sheet("7. Sign-off")
    title(ws, "QA sign-off", 2)
    head(ws, 3, ["Item", "Detail"], [34, 76])
    body(ws, 4, [
        ["Recommendation", verdict],
        ["Rationale", rationale],
        ["Validation report", Path(data["report_path"]).name],
        ["Defect log", Path(data["defect_path"]).name if data["defect_path"] else "n/a"],
        ["Execution log", meta.get("log_file", "")],
        ["Evidence location", "reports/evidence/"],
        ["Prepared by", "Parashurama Nagalapurada — Senior Associate, Quality Assurance"],
        ["Date", datetime.now().strftime("%d-%b-%Y")],
        ["Reviewed by", ""],
        ["Review date", ""],
    ], bold_first=True)

    out = Path(out_path or REPORT_DIR /
               f"Final_QA_Summary_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    wb.save(out)
    return out, verdict, high, blocked


def main():
    ap = argparse.ArgumentParser(description="Generate the Final QA Summary")
    ap.add_argument("--report", help="Path to validation_report_*.xlsx")
    ap.add_argument("--defects", help="Path to defect_log_*.xlsx")
    ap.add_argument("--out", help="Output path")
    ap.add_argument("--auto", action="store_true",
                    help="Use the most recent report and defect log in reports/")
    a = ap.parse_args()

    data = load_inputs(a.report, a.defects)
    out, verdict, high, blocked = build(data, a.out)

    print("\n" + "=" * 74)
    print("  FINAL QA SUMMARY GENERATED")
    print("=" * 74)
    print(f"  Source report : {Path(data['report_path']).name}")
    print(f"  Source defects: {Path(data['defect_path']).name if data['defect_path'] else 'n/a'}")
    print(f"  Output        : {out}")
    print(f"  Recommendation: {verdict}")
    print(f"  High-severity : {high}   Blocked: {blocked}")
    print("-" * 74)
    print("  Review sheets 3-6 before submitting. The recommendation is derived")
    print("  from thresholds in THRESHOLDS and reflects QA judgement you own.")
    print("=" * 74 + "\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
